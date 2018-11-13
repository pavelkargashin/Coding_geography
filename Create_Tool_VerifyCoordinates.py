# coding: utf8
import openpyxl, re, shutil

# sheet extraction
def sheet_extraction(xlsx):
    input_xslx = openpyxl.load_workbook(xslx)
    sheets = input_xslx.sheetnames
    for item in sheets:
        sheet = input_xslx[item]
        parameters = searching_xslx_coordinates(sheet)
    return sheet, parameters, input_xslx


# Searching columns with coordinates
def searching_xslx_coordinates(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Longitude":
                lon_index = row.index(cell)
            if cell.value == "Latitude":
                lat_index = row.index(cell)
    lon_length = tuple(sheet.columns)[lon_index].__len__()
    return lon_index, lat_index, lon_length


# Extracting coordinate values from xlsx
def reading_xslx_values(sheet, lon_index, lat_index, lon_length):
    xy_list = []
    for i in range(1, lon_length):
        text = []
        lon = tuple(sheet.columns)[lon_index][i].value
        text.append(str(lon))
        lat = tuple(sheet.columns)[lat_index][i].value
        text.append(str(lat))
        xy_list.append(text)
    return xy_list


# Writing coordinate values to xlsx (n = 0 or 1, means main and alternative pair of coordinate accordingly)
def writing_xslx_values(xy_list, sheet, lon_index, lat_index, lon_length, n):
    for i in range(1, lon_length):
        sheet.cell(row=i+1, column=lon_index+1).value = xy_list[i-1][n][0]
        sheet.cell(row=i+1, column=lat_index+1).value = xy_list[i-1][n][1]


# Transforming decimal degrees like хх.хх''хх''.хх'' и хх.хх''хх.хх'' in decimal degrees
def decimating_pseudoDDD(input):
    result0 = re.sub(r"''\.", "''", input)
    result1 = re.split(r"[']+", result0)
    Base_Value = float(result1[0]) + float(result1[1]) / 10000
    try:
        Value = Base_Value + float(result1[2])/1000000
    except:
        Value = Base_Value
    return Value


# transforming degrees-minutes-seconds like хх.хх''хх''.хх'' and хх.хх''хх.хх'' in decimal degrees
def decimating_pseudoDMS(input):
    result0 = re.sub(r"''\.", ".", input)
    result = re.split(r"['',\.]+", result0)
    if float(result[1]) < 60 or float(result[2]) < 60:
        Value = float(result[0]) + float(result[1])/60 + float(result[2]+"." +result[3])/3600
    else:
        Value = 0
    return Value


# transforming degrees-minutes-seconds like хх°хх'хх'' in decimal degrees
def decimating_DMS(input):
    result0 = re.sub(r',', '.', input)
    result00 = re.sub(r"[°◦']+", 'o', result0)
    result = re.split(r'o', result00)
    Value = float(result[0]) + float(result[1])/60 + float(result[2])/3600
    return Value


# transforming degrees-minutes-seconds like xx.xxxxxx in decimal degrees
def decimating_falseDDD(input):
    result = re.split(r"[\.]", input)
    m = float(result[1][0:2])
    tale = result[1][2:]
    power = len(tale)-2
    s = float(tale)/(10**power)
    value = float(result[0]) + m/60 + s/3600
    return value


# decision tree for coordinates reading
def Conversion(Text):
    Output = []
    Output_Alternative = []
    # if points have no coordinates
    if Text[0] == Text[1] == "99999999":
        Output.append(float(9))
        Output.append(float(116))
    else:
        Input = str(Text[0]) + "_" + str(Text[1])
        result1 = re.findall(r"[°◦]", Input)
        # if coordinates have symbol "°"
        if len(result1) is not 0:
            for Value in Text:
                Output.append(decimating_DMS(Value))
        else:
            result2 = re.findall(r"''", Input)
            # if coordinates have symbol "''"
            if len(result2) is not 0:
                result3 = re.findall(r"\d'\d", Input)
                # if coordinates have symbol "'"
                if len(result3) is not 0:
                    for Value in Text:
                        Output.append(decimating_pseudoDMS(Value))
                # if coordinates have symbol "'" but minutes or seconds are more than 60
                elif decimating_pseudoDMS(Text[0]) == 0 or decimating_pseudoDMS(Text[1]) == 0:
                    for Value in Text:
                        Output.append(decimating_pseudoDDD(Value))
                else:
                    # reading coordinates with two different functions
                    for Value in Text:
                        Output.append(decimating_pseudoDMS(Value))
                        Output_Alternative.append(decimating_pseudoDDD(Value))
                    # appending key value 1 to the output representing doubts in reading coordinates
                    Output.append(1)
                    Output_Alternative.append(1)
            # if coordinates are likely to be truly decimal
            else:
                # reading coordinates with two different functions
                for Value in Text:
                    Output.append(float(Value))
                    Output_Alternative.append(decimating_falseDDD(Value))
                # appending key value 1 to the output representing doubts in reading coordinates
                Output.append(1)
                Output_Alternative.append(1)
    # appending key value 0 to the output representing no doubts in reading coordinates
    if len(Output) == 2:
        Output.append(0)
    # don't forget that we are in the southern hemisphere
    if Output[1] > 0:
        Output[1] = Output[1]*(-1)
    if len(Output_Alternative) > 0 and Output_Alternative[1] > 0:
        Output_Alternative[1] = Output_Alternative[1]*(-1)
    return Output, Output_Alternative


xslx = r'D:\YandexDisk\Projects\Bali\AirSungai.xlsx'
xslx_verified = shutil.copyfile(xslx, r'D:\YandexDisk\Projects\Bali\AirSungai_Verified.xlsx')
xslx_alternative = shutil.copyfile(xslx, r'D:\YandexDisk\Projects\Bali\AirSungai_Alternative.xlsx')
result = sheet_extraction(xslx)
list = reading_xslx_values(result[0],result[1][0], result[1][1], result[1][2])
xy_list = []
for text in list:
    xy = Conversion(text)
    xy_list.append(xy)
result_verified = sheet_extraction(xslx_verified)
result_alternative = sheet_extraction(xslx_alternative)
print xy_list
writing_xslx_values(xy_list, result_verified[0], result_verified[1][0], result_verified[1][1], result_verified[1][2], 0)
writing_xslx_values(xy_list, result_alternative[0], result_alternative[1][0], result_alternative[1][1], result_alternative[1][2], 1)
result_verified[2].save(xslx_verified)
result_alternative[2].save(xslx_verified)
