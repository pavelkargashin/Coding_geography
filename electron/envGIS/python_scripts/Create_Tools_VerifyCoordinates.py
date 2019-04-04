# coding: utf8
import openpyxl, re, shutil, sys
reload(sys)
sys.setdefaultencoding('utf8')

# Transforming decimal degrees like хх.хх''хх''.хх'' и хх.хх''хх.хх'' in decimal degrees
def decimating_pseudoDDD(input):
    result0 = re.sub(r"''\.", "''", input)
    result1 = re.split(r"[']+", result0)
    Base_Value = float(result1[0]) + float(result1[1]) / 10000
    try:
        Value = Base_Value + float(result1[2])/1000000
    except:
        Value = Base_Value
    return Value


# transforming degrees-minutes-seconds like хх.хх''хх''.хх'' and хх.хх''хх.хх'' in decimal degrees
def decimating_pseudoDMS(input):
    result0 = re.sub(r"''\.", ".", input)
    result = re.split(r"['',\.]+", result0)
    if float(result[1]) < 60 or float(result[2]) < 60:
        Value = float(result[0]) + float(result[1])/60 + float(result[2]+"." +result[3])/3600
    else:
        Value = 0
    return Value


# transforming degrees-minutes-seconds like хх°хх'хх'' in decimal degrees
def decimating_DMS(input):
    result0 = re.sub(r',', '.', input)
    result00 = re.sub(r"[°◦']+", 'o', result0)
    result = re.split(r'o', result00)
    value = float(result[0]) + float(result[1])/60 + float(result[2])/3600
    return value


# transforming degrees-minutes-seconds like xx.xxxxxx in decimal degrees
def decimating_falseDDD(input):
    result = re.split(r"[\.]", input)
    print result
    m = float(result[1][0:2])
    tale = result[1][2:]
    power = len(tale)-2
    s = float(tale)/(10**power)
    value = float(result[0]) + m/60 + s/3600
    return value


# decision tree for coordinates reading
def conversion(text):
    output = []
    output_alternative = []
    # if points have no coordinates
    if text[0] == text[1] == "99999999":
        output.append(float(116))
        output.append(float(-9))
    else:
        Input = str(text[0]) + "_" + str(text[1])
        result1 = re.findall(r"[°◦]", Input)
        # if coordinates have symbol "°"
        if len(result1) is not 0:
            for value in text:
                output.append(decimating_DMS(value))
        else:
            result2 = re.findall(r"''", Input)
            # if coordinates have symbol "''"
            if len(result2) is not 0:
                result3 = re.findall(r"\d'\d", Input)
                # if coordinates have symbol "'"
                if len(result3) is not 0:
                    for value in text:
                        output.append(decimating_pseudoDMS(value))
                # if coordinates have symbol "'" but minutes or seconds are more than 60
                elif decimating_pseudoDMS(text[0]) == 0 or decimating_pseudoDMS(text[1]) == 0:
                    for value in text:
                        output.append(decimating_pseudoDDD(value))
                else:
                    # reading coordinates with two different functions
                    for value in text:
                        output.append(decimating_pseudoDMS(value))
                        output_alternative.append(decimating_pseudoDDD(value))
            # if coordinates are likely to be truly decimal
            else:
                # reading coordinates with two different functions
                for value in text:
                    output.append(float(value))
                    output_alternative.append(decimating_falseDDD(value))
    # don't forget that we are in the southern hemisphere
    if output[1] > 0:
        output[1] = output[1]*(-1)
    if len(output_alternative) > 0 and output_alternative[1] > 0:
        output_alternative[1] = output_alternative[1] * (-1)
    return output, output_alternative


# sheet extraction
def sheet_extraction(xlsx):
    input_xslx = openpyxl.load_workbook(xlsx)
    sheet_names = input_xslx.sheetnames
    sheets = []
    for item in sheet_names:
        sheet = input_xslx[item]
        sheets.append(sheet)
    return sheets


# Searching columns with coordinates
def searching_xslx_coordinates(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Longitude":
                lon_index = row.index(cell)
            if cell.value == "Latitude":
                lat_index = row.index(cell)
    lon_length = tuple(sheet.columns)[lon_index].__len__()
    return lon_index, lat_index, lon_length


# Extracting coordinate values from xlsx
def reading_xslx_values(sheet, lon_index, lat_index, lon_length):
    xy_list = []
    key_list = []
    for i in range(1, lon_length):
        text = []
        if tuple(sheet.columns)[lon_index][i].value is not None:
            lon = tuple(sheet.columns)[lon_index][i].value
            text.append(str(lon))
            lat = tuple(sheet.columns)[lat_index][i].value
            text.append(str(lat))
            # Sequence number for row
            key_list.append(i)
            xy_list.append(text)
    dictionary = dict(zip(key_list, xy_list))
    return dictionary


# Writing coordinate values to xlsx (n = 0 or 1, means main and alternative output pair of coordinates accordingly)
def writing_xslx_values(sheet, dictionary, lon_index, lat_index, workbook):
    input_xlsx = openpyxl.load_workbook(workbook)
    print workbook
    if workbook == xslx_verified:
        n = 0
    else:
        n = 1
    dict_xy_list = {}
    print dictionary
    for key in dictionary:
        print key
        print dictionary[key]
        xy = conversion(dictionary[key])
        print xy
        #adding recalculated coordinates to the new dictionary
        dict_xy_list[key] = xy
        if len(dict_xy_list[key][n]) == 2:
            sheet.cell(row=key+1, column=lon_index+1).value = dict_xy_list[key][n][0]
            sheet.cell(row=key+1, column=lat_index+1).value = dict_xy_list[key][n][1]
            print sheet.cell(row=key + 1, column=lon_index + 1).value
        else:
            print "delete row"
            sheet.delete_rows(key, 1)
        if n == 1 and key != 0:
            if key not in dict_xy_list.keys():
                print "No alternative"
                sheet.delete_rows(key, 1)


#input = raw_input()
#xslx = "r'" + input + "'"
xslx = r'D:\YandexDisk\Projects\Bali\Copy of  StatData180201_Air_Sumur.xlsx'
xslx_verified = xslx[:-5] + r'_Verified.xlsx'
xslx_alternative = xslx[:-5] + r'_Alternative.xlsx'
output_files = [xslx_verified, xslx_alternative]
for item in output_files:
    shutil.copyfile(xslx, item)
    input_xlsx = openpyxl.load_workbook(item)
    print str(input_xlsx)
    sheets = sheet_extraction(item)
    for sheet in list(sheets):
        res = searching_xslx_coordinates(sheet)
        lon_index = res[0]
        lat_index = res[1]
        lon_length = res[2]
        dictionary = reading_xslx_values(sheet, lon_index,lat_index, lon_length)
        writing_xslx_values(sheet, dictionary, lon_index, lat_index, item)
    print "Save " + str(input_xlsx) + " with name " + str(item)
    input_xlsx.save(item)


