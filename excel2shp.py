#-*-coding:utf-8-*-

import csv
import openpyxl
import parameters
import ogr
import osr
import re


TestDataFolder = parameters.TestDataFolder
workData = parameters.workData
exceldata = TestDataFolder+workData


# Функция для считывания файла ехсеl и выдачи списка листов. Список дистов записывается в файл txt
def explore_workbook(inputexcel):
    myExcel = openpyxl.load_workbook(inputexcel)
    mySheets = myExcel.sheetnames # get all sheets via list
    myDoc = open(TestDataFolder + 'TEST.txt', 'w')
    for item in mySheets:
        myDoc.write(item+'\n')
    myDoc.close()
    return mySheets


def choose_sheets(sheet_list):
    my_sheet_list = []
    for item in sheet_list:
        search_result = re.search(r'Kualitas Air', item)
        if search_result != None:
            my_sheet_list.append(item)
        else:
            continue
    return my_sheet_list


def form_names(list_of_sheets):
    naming_dictionary = {}
    for item in list_of_sheets:
        dic_value = re.split(r' ', item)
        naming_dictionary[item] = dic_value[2]+dic_value[3]
    return naming_dictionary


def export_data_to_dictionary(inputexcel, currentSheet, attr_names):
    myTable = []
    list_temp = []
    myExcel = openpyxl.load_workbook(inputexcel)
    mySheet = myExcel[currentSheet]
    for row_val in range(2, mySheet.max_row):
        for col_val in range(1,mySheet.max_column):
            list_temp.append(mySheet.cell(column = col_val, row = row_val).value)
        myRow = dict(zip(attr_names,list_temp))
        myTable.append(myRow)
        list_temp=[]
        myRow={}
    return myTable


# Экспорт заголовков колонок таблицы в список.
def create_fields_list(inputexcel, currentSheet):
    list_temp = []
    myExcel = openpyxl.load_workbook(inputexcel)
    mySheet = myExcel[currentSheet]
    row_val=1
    for col_val in range(1, mySheet.max_column):
        # print (mySheet.cell(column=col_val, row=row_val).value)
        list_temp.append(mySheet.cell(column=col_val, row=1).value)
    return list_temp

#Функция для обрезки названий полей до 10 симоволов. Ограничение из-за шейпфайла
def shorten_field_name(field_list):
    new_list = []
    for item in field_list:
        # print item
        if len(item)>10:
            tempVal = item[0:10]
            # print tempVal
            new_list.append(tempVal)
        else:
            new_list.append(item)
    return new_list


def create_shapefile(ShapeName, column_names, inputTable):
    driver  = ogr.GetDriverByName("ESRI Shapefile")
    datasource = driver.CreateDataSource(ShapeName)
    srs = osr.SpatialReference()
    srs.ImportFromEPSG(4326)
    layer = datasource.CreateLayer('test', srs, ogr.wkbPoint)
    i=1
    for item in column_names:
        # print 'Column name is: ',item
        fieldname = 'field'+str(i)
        if item == 'RiverName' or item=='NameLocati' or item =='Stage' or item == "Monitoring":
           fieldname = ogr.FieldDefn(str(item), ogr.OFTString)
           fieldname.SetWidth(50)
           layer.CreateField(fieldname)
           i+=1
        else:
           layer.CreateField(ogr.FieldDefn(str(item), ogr.OFTReal))
    for row in inputTable:
        feature = ogr.Feature(layer.GetLayerDefn())
        for name in column_names:
            feature.SetField(str(name), row[str(name)])
        wkt="POINT(%f %f)" % (float(row['KoordinatL']), float(row['KoordinatB']))
        point = ogr.CreateGeometryFromWkt(wkt)
        feature.SetGeometry(point)
        layer.CreateFeature(feature)
        feature = None
    datasource = None





# ФУНКЦИИ ПОКА НЕ ИСПОЛЬЗУЮТСЯ!!!!


# Функция для экспорта данных в csv формат. На входе нужен файл и наименование листа.
# Экспортируются данные, начиная со второй строки, то есть заголовки не выходят
# Проблемы с экспортом - кодировки и т.д.!!!!!!!!!!!!!!!!!
def export_data_to_csv(inputexcel, currentSheet):
    list_temp = []
    myExcel = openpyxl.load_workbook(inputexcel)
    mySheet = myExcel[currentSheet]
    with open(TestDataFolder+"new_csv.csv", 'wb') as mycsv:
        fw = csv.writer(mycsv, delimiter = ',')
        for row_val in range(2, mySheet.max_row+1):
            for col_val in range(1,mySheet.max_column):
                list_temp.append(mySheet.cell(column = col_val, row = row_val).value)
            # print(list_temp)
            fw.writerow(list_temp)
            list_temp=[]
    mycsv.close()


# With csv writing
def create_fields_list2(inputexcel, currentSheet):
    list_temp = []
    myExcel = openpyxl.load_workbook(inputexcel)
    mySheet = myExcel[currentSheet]
    row_val=1
    with open(TestDataFolder+'attrNames.csv', 'w') as curCsv:
        MyWriter = csv.writer(curCsv, delimiter = ',')

        for col_val in range(1, mySheet.max_column+1):
            # print (mySheet.cell(column=col_val, row=row_val).value)
            list_temp.append(mySheet.cell(column=col_val, row=1).value)
        MyWriter.writerow(list_temp)
        curCsv.close()

    return list_temp