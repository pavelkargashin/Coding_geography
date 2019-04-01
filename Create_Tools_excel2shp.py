#-*-coding:utf-8-*-


import openpyxl
import ogr
import osr
import re


# Функция для считывания файла ехсеl и выдачи списка листов. Список листов записывается в файл txt
def explore_workbook(inputDataFolder, excelName):
    myExcel = openpyxl.load_workbook(inputDataFolder+excelName)
    mySheets = myExcel.sheetnames # get all sheets via list
    myDoc = open(inputDataFolder + 'TEST.txt', 'w')
    for item in mySheets:
        myDoc.write(item+'\n')
    myDoc.close()
    return mySheets


def choose_sheets(sheet_list, keyword4search):
    my_sheet_list = []
    for item in sheet_list:
        search_result = re.search(keyword4search, item)
        if search_result != None:
            my_sheet_list.append(item)
        else:
            continue
    return my_sheet_list


def form_names(list_of_sheets):
    naming_dictionary = {}
    for item in list_of_sheets:
        naming_dictionary[item] = str(item)
    return naming_dictionary


def export_data_to_dictionary(inputexcel, currentSheet, attr_names):
    myTable = []
    list_temp = []
    myExcel = openpyxl.load_workbook(inputexcel)
    mySheet = myExcel[currentSheet]
    for row_val in range(2, mySheet.max_row):
        for col_val in range(1,mySheet.max_column+1):
            list_temp.append(mySheet.cell(column = col_val, row = row_val).value)
        myRow = dict(zip(attr_names,list_temp))
        myTable.append(myRow)
        list_temp = []
        myRow = {}
    return myTable


# Экспорт заголовков колонок таблицы в список.
def create_fields_list(inputexcel, currentSheet):
    list_temp = []
    myExcel = openpyxl.load_workbook(inputexcel)
    mySheet = myExcel[currentSheet]
    for col_val in range(1, mySheet.max_column+1):
        list_temp.append(mySheet.cell(column=col_val, row=1).value)
        print list_temp
    return list_temp

#Функция для обрезки названий полей до 10 симоволов. Ограничение из-за шейпфайла
def shorten_field_name(field_list):
    new_list = []
    for item in field_list:
        print str(item)
        if len(item)>10:
            tempVal = item[0:10]
            new_list.append(tempVal)
        else:
            new_list.append(item)
    return new_list


def create_shapefile(ShapeName, column_names, inputTable):
    driver  = ogr.GetDriverByName("ESRI Shapefile")
    datasource = driver.CreateDataSource(ShapeName)
    srs = osr.SpatialReference()
    srs.ImportFromEPSG(4326)
    layer = datasource.CreateLayer('test', srs, ogr.wkbPoint)
    i=1
    layer.CreateField(ogr.FieldDefn("LoadID", ogr.OFTReal))
    for item in column_names:
        fieldname = 'field'+str(i)
        textfields = ('RiverName', 'Location', 'Stage', 'Longitude', 'Latitude', 'LakeName', 'Area', 'Point', 'Basin')
        if item in textfields:
           fieldname = ogr.FieldDefn(str(item), ogr.OFTString)
           fieldname.SetWidth(50)
           layer.CreateField(fieldname)
           i+=1
        else:
           layer.CreateField(ogr.FieldDefn(str(item), ogr.OFTReal))
    tempi = 1
    for row in inputTable:
        if row['Year']!=None:
            # Создание точек с атрибутами
            inputcoord = []
            inputcoord.append(str(row['Longitude']))
            inputcoord.append(str(row['Latitude']))
            wkt = "POINT(%f %f)" % (float(inputcoord[0]), float(inputcoord[1]))
            point = ogr.CreateGeometryFromWkt(wkt)
            feature = ogr.Feature(layer.GetLayerDefn())
            feature.SetGeometry(point)
            feature.SetField("LoadID", tempi)
            for name in column_names:
                feature.SetField(str(name), row[str(name)])
            layer.CreateFeature(feature)
            feature = None
            tempi += 1
        else:
            continue

    datasource = None


        # outputcoord = Coordinates.Conversion(inputcoord)
        #     if len(outputcoord[1])==0:
        #         wkt="POINT(%f %f)" % (float(outputcoord[0][1]), float(outputcoord[0][0]))
        #         point = ogr.CreateGeometryFromWkt(wkt)
        #         feature = ogr.Feature(layer.GetLayerDefn())
        #         feature.SetGeometry(point)
        #         feature.SetField("LoadID", tempi)
        #         # feature.SetField("Doubt", int(outputcoord[0][2]))
        #         for name in column_names:
        #             feature.SetField(str(name), row[str(name)])
        #         layer.CreateFeature(feature)
        #         feature = None
        #     elif len(outputcoord[1])>0:
        #         for item in outputcoord:
        #             wkt = "POINT(%f %f)" % (float(item[1]), float(item[0]))
        #             point = ogr.CreateGeometryFromWkt(wkt)
        #             feature = ogr.Feature(layer.GetLayerDefn())
        #             feature.SetGeometry(point)
        #             feature.SetField("LoadID", tempi)
        #             # feature.SetField("Doubt", int(outputcoord[1][2]))
        #             for name in column_names:
        #                 feature.SetField(str(name), row[str(name)])
        #             layer.CreateFeature(feature)
        #             feature = None
        #     tempi+=1
        # else:
        #     continue

